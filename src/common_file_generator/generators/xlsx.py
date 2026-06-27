"""Complexity-driven Excel workbook generator.

Builds an ``.xlsx`` from scratch (not template injection). A workbook is a set of
**sheets**; each sheet holds one or more titled data tables. The complexity level
decides what each sheet contains: a plain table at ``minimal``, then formula
total rows, charts, styling, and multiple tables at higher levels.

Cell data is typed (numbers, text via the seeded ``Lorem``, and dates) and totals
are written as formula strings (``=SUM(...)`` / ``=AVERAGE(...)``) that the
spreadsheet app evaluates on open. The same ``(complexity, sheets, seed)`` always
yields the identical workbook.
"""

from __future__ import annotations

import datetime
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from common_file_generator.core.complexity import Complexity, sheet_feature_pool
from common_file_generator.core.lorem import Lorem

_HEADER_FILL = "2E86C1"
_CHART_KINDS = ("bar", "line", "pie")
_EPOCH = datetime.date(2020, 1, 1)


class XlsxComplexityGenerator:
    """Generate an Excel workbook at a chosen complexity level."""

    def __init__(
        self,
        complexity: Complexity = Complexity.STANDARD,
        *,
        sheets: int = 3,
        seed: int = 0,
        rows: int | None = None,
        cols: int | None = None,
    ) -> None:
        if sheets < 1:
            raise ValueError("sheets must be at least 1")
        if rows is not None and rows < 1:
            raise ValueError("rows must be at least 1")
        if cols is not None and cols < 3:
            raise ValueError("cols must be at least 3 (label, value, date)")
        self.complexity = complexity
        self.sheets = sheets
        self.features = sheet_feature_pool(complexity)
        self.rows_per_table = rows if rows is not None else self.features.rows_per_table
        self._cols = cols
        self._rng = random.Random(seed)
        self._lorem = Lorem(self._rng)

    def build(self) -> Workbook:
        workbook = Workbook()
        default = workbook.active
        for index in range(1, self.sheets + 1):
            sheet = default if index == 1 else workbook.create_sheet()
            sheet.title = f"{index} {self._lorem.title(2)}"[:31]
            self._fill_sheet(sheet)
        return workbook

    def save(self, out_path: str | Path) -> Path:
        out_path = Path(out_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        self.build().save(str(out_path))
        return out_path

    # --- sheet content --------------------------------------------------

    def _fill_sheet(self, sheet: Worksheet) -> None:
        top = 1
        for _ in range(self.features.tables_per_sheet):
            top = self._add_table(sheet, top) + 2

    def _add_table(self, sheet: Worksheet, top: int) -> int:
        """Write a titled table starting at row ``top``; return its last row."""
        cols = self._cols if self._cols is not None else self._rng.randint(3, 5)
        rows = self.rows_per_table

        title_cell = sheet.cell(row=top, column=1, value=self._lorem.title(3))
        title_cell.font = Font(bold=True, size=13)
        header_row = top + 1
        data_start = header_row + 1
        data_end = data_start + rows - 1

        # First column is a label, last column is a date, the middle are numbers.
        date_col = cols
        headers = ["Item"]
        headers += [self._lorem.title(1) for _ in range(cols - 2)]
        headers += ["Date"]
        for c, text in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=c, value=text)
            if self.features.styling:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)

        for r in range(data_start, data_end + 1):
            sheet.cell(row=r, column=1, value=self._lorem.title(2))
            for c in range(2, cols + 1):
                if c == date_col:
                    days = self._rng.randint(0, 1460)
                    cell = sheet.cell(
                        row=r, column=c, value=_EPOCH + datetime.timedelta(days=days)
                    )
                    cell.number_format = "yyyy-mm-dd"
                else:
                    cell = sheet.cell(row=r, column=c, value=self._rng.randint(1, 1000))
                    if self.features.styling:
                        cell.number_format = "#,##0"

        # Numeric columns are 2 .. cols-1 (last column is the date).
        numeric_end = cols - 1
        last_row = data_end
        if self.features.formulas:
            last_row = self._add_totals(sheet, data_start, data_end, numeric_end)
        if self.features.styling:
            self._style_widths(sheet, cols)
        if self.features.charts and numeric_end >= 2:
            self._add_chart(sheet, header_row, data_start, data_end, numeric_end)
        return last_row

    def _add_totals(
        self, sheet: Worksheet, data_start: int, data_end: int, numeric_end: int
    ) -> int:
        total_row = data_end + 1
        sheet.cell(row=total_row, column=1, value="Total")
        for c in range(2, numeric_end + 1):
            letter = get_column_letter(c)
            func = "SUM" if c % 2 == 0 else "AVERAGE"
            sheet.cell(
                row=total_row,
                column=c,
                value=f"={func}({letter}{data_start}:{letter}{data_end})",
            )
        return total_row

    def _style_widths(self, sheet: Worksheet, cols: int) -> None:
        sheet.column_dimensions["A"].width = 22
        for c in range(2, cols + 1):
            sheet.column_dimensions[get_column_letter(c)].width = 12

    def _add_chart(
        self,
        sheet: Worksheet,
        header_row: int,
        data_start: int,
        data_end: int,
        numeric_end: int,
    ) -> None:
        kind = self._rng.choice(_CHART_KINDS)
        chart = {"bar": BarChart, "line": LineChart, "pie": PieChart}[kind]()
        chart.title = self._lorem.title(2)
        values = Reference(
            sheet, min_col=2, min_row=header_row, max_col=numeric_end, max_row=data_end
        )
        categories = Reference(sheet, min_col=1, min_row=data_start, max_row=data_end)
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(categories)
        # Anchor to the right of the data + date column.
        anchor = f"{get_column_letter(numeric_end + 3)}{header_row}"
        sheet.add_chart(chart, anchor)
