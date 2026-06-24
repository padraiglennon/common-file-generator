"""Excel injector (openpyxl).

Excel's natural named tag is the **defined name** (Formulas > Name Manager).
A non-technical user names cells/anchors in the workbook, and the config refers
to those names:

* ``text``   - a defined name pointing at a single cell receives the value.
* ``tables`` - a defined name marks the top-left anchor; rows fill down/right
               from there (existing cells only; the sheet is not extended).
* ``media``  - a defined name marks the cell an image is anchored to.

No sheets, rows, or columns are created; only existing cells are written.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ms_office_file_generator.core.config import Config, MediaSpec, TableSpec
from ms_office_file_generator.core.injector import Injector


class XlsxInjector(Injector):
    """Inject data into an Excel Golden template."""

    extensions = frozenset({".xlsx"})

    def _load(self, path: Path) -> object:
        return load_workbook(str(path))

    def _save(self, path: Path) -> None:
        self._document.save(str(path))

    def _anchor(self, name: str) -> tuple[Worksheet, int, int] | None:
        """Resolve a defined name to (worksheet, row, col), or None."""
        defined = self._document.defined_names
        if name not in defined:
            return None
        destinations = list(defined[name].destinations)
        if not destinations:
            return None
        sheet_name, ref = destinations[0]
        ref = ref.replace("$", "")
        worksheet = self._document[sheet_name]
        row, col = coordinate_to_tuple(ref)
        return worksheet, row, col

    # --- text -----------------------------------------------------------

    def _inject_text(self, config: Config) -> None:
        for name, value in config.text.items():
            anchor = self._anchor(name)
            if anchor is None:
                self.report.warning(
                    f"text.{name}",
                    "no defined name with this name was found in the workbook; "
                    "this value was not used.",
                )
                continue
            worksheet, row, col = anchor
            worksheet.cell(row=row, column=col, value=value)

    # --- tables ---------------------------------------------------------

    def _inject_tables(self, config: Config) -> None:
        for spec in config.tables:
            anchor = self._anchor(spec.name)
            if anchor is None:
                self.report.error(
                    f"tables.{spec.name}",
                    "no defined name with this name was found in the workbook.",
                )
                continue
            self._fill_table(spec, *anchor)

    def _fill_table(
        self, spec: TableSpec, worksheet: Worksheet, top: int, left: int
    ) -> None:
        data_rows = list(spec.rows)
        if spec.header is not None:
            data_rows = [spec.header, *data_rows]

        max_row, max_col = worksheet.max_row, worksheet.max_column
        avail_rows = max_row - top + 1
        avail_cols = max_col - left + 1

        if len(data_rows) > avail_rows:
            self.report.warning(
                f"tables.{spec.name}",
                f"you provided {len(data_rows)} rows but only {avail_rows} fit "
                f"below the '{spec.name}' anchor; kept the first {avail_rows}.",
            )
        for r, row in enumerate(data_rows[:avail_rows]):
            if len(row) > avail_cols:
                self.report.warning(
                    f"tables.{spec.name}",
                    f"row {r} has {len(row)} cells but only {avail_cols} columns "
                    f"fit from the anchor; kept the first {avail_cols}.",
                )
            for c, value in enumerate(row[:avail_cols]):
                worksheet.cell(row=top + r, column=left + c, value=value)

    # --- media ----------------------------------------------------------

    def _inject_media(self, config: Config) -> None:
        for spec in config.media:
            self._place_media(spec)

    def _place_media(self, spec: MediaSpec) -> None:
        anchor = self._anchor(spec.name)
        if anchor is None:
            self.report.error(
                f"media.{spec.name}",
                "no defined name with this name was found in the workbook.",
            )
            return
        if not spec.path.is_file():
            self.report.error(
                f"media.{spec.name}",
                f"image file not found: {spec.path}.",
            )
            return
        worksheet, row, col = anchor
        image = XLImage(str(spec.path))
        worksheet.add_image(image, f"{get_column_letter(col)}{row}")
