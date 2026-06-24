"""Tests for the complexity-driven Excel workbook generator."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from ms_office_file_generator.core import Complexity, generate_sheet
from ms_office_file_generator.generators import XlsxComplexityGenerator


def _formulas(path: Path) -> list[str]:
    wb = load_workbook(str(path))
    return [
        cell.value
        for name in wb.sheetnames
        for row in wb[name].iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]


def _chart_count(path: Path) -> int:
    wb = load_workbook(str(path))
    return sum(len(wb[name]._charts) for name in wb.sheetnames)


def test_sheet_count_matches_request(tmp_path: Path) -> None:
    out = tmp_path / "w.xlsx"
    generate_sheet(str(out), complexity="standard", sheets=4, seed=1)
    assert len(load_workbook(str(out)).sheetnames) == 4


def test_invalid_sheet_count_raises() -> None:
    with pytest.raises(ValueError, match="sheets must be at least 1"):
        XlsxComplexityGenerator(Complexity.MINIMAL, sheets=0)


def test_invalid_rows_raises() -> None:
    with pytest.raises(ValueError, match="rows must be at least 1"):
        XlsxComplexityGenerator(Complexity.MINIMAL, sheets=1, rows=0)


def test_minimal_has_no_formulas_or_charts(tmp_path: Path) -> None:
    out = tmp_path / "min.xlsx"
    generate_sheet(str(out), complexity="minimal", sheets=3, seed=3)
    assert _formulas(out) == []
    assert _chart_count(out) == 0


def test_maximum_has_formulas_and_charts(tmp_path: Path) -> None:
    out = tmp_path / "max.xlsx"
    generate_sheet(str(out), complexity="maximum", sheets=3, seed=7)
    formulas = _formulas(out)
    assert any(f.startswith("=SUM") for f in formulas)
    assert any(f.startswith("=AVERAGE") for f in formulas)
    assert _chart_count(out) > 0


def test_seed_is_deterministic(tmp_path: Path) -> None:
    a, b = tmp_path / "a.xlsx", tmp_path / "b.xlsx"
    generate_sheet(str(a), complexity="complex", sheets=3, seed=42)
    generate_sheet(str(b), complexity="complex", sheets=3, seed=42)
    assert _cells(a) == _cells(b)


def test_different_seed_differs(tmp_path: Path) -> None:
    a, b = tmp_path / "a.xlsx", tmp_path / "b.xlsx"
    generate_sheet(str(a), complexity="complex", sheets=3, seed=1)
    generate_sheet(str(b), complexity="complex", sheets=3, seed=2)
    assert _cells(a) != _cells(b)


def test_cols_override(tmp_path: Path) -> None:
    out = tmp_path / "w.xlsx"
    generate_sheet(str(out), complexity="minimal", sheets=1, seed=1, cols=4)
    ws = load_workbook(str(out)).active
    # Header is on row 2; 4 columns means 4 non-empty header cells.
    headers = [ws.cell(row=2, column=c).value for c in range(1, 5)]
    assert all(headers)
    assert ws.cell(row=2, column=5).value is None


def test_invalid_cols_raises() -> None:
    with pytest.raises(ValueError, match="cols must be at least 3"):
        XlsxComplexityGenerator(Complexity.MINIMAL, sheets=1, cols=2)


def test_styled_numbers_are_formatted(tmp_path: Path) -> None:
    out = tmp_path / "w.xlsx"
    generate_sheet(str(out), complexity="standard", sheets=1, seed=1)
    ws = load_workbook(str(out)).active
    # A numeric data cell (row 3, col 2) carries the thousands format.
    assert ws.cell(row=3, column=2).number_format == "#,##0"


def test_maximum_is_denser_than_complex(tmp_path: Path) -> None:
    cx = tmp_path / "cx.xlsx"
    mx = tmp_path / "mx.xlsx"
    generate_sheet(str(cx), complexity="complex", sheets=1, seed=1)
    generate_sheet(str(mx), complexity="maximum", sheets=1, seed=1)
    # maximum has 3 tables per sheet vs complex's 2, so more total rows.
    assert load_workbook(str(mx)).active.max_row > load_workbook(str(cx)).active.max_row


def test_rows_override(tmp_path: Path) -> None:
    out = tmp_path / "w.xlsx"
    generate_sheet(str(out), complexity="minimal", sheets=1, seed=1, rows=20)
    ws = load_workbook(str(out)).active
    # title + header + 20 data rows = 22.
    assert ws.max_row == 22


def _cells(path: Path) -> list:
    wb = load_workbook(str(path))
    return [
        (name, cell.coordinate, str(cell.value))
        for name in wb.sheetnames
        for row in wb[name].iter_rows()
        for cell in row
    ]
