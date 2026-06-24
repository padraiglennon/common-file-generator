"""Tests for the thin CLI wrapper."""

from __future__ import annotations

import json
from pathlib import Path

from pptx import Presentation

from ms_office_file_generator.cli import main


def _config_file(tmp_path: Path, data: dict) -> Path:
    path = tmp_path / "config.json"
    path.write_text(json.dumps(data), encoding="utf-8")
    return path


def test_fill_success(pptx_template: Path, tmp_path: Path) -> None:
    out = tmp_path / "out.pptx"
    code = main(
        [
            "fill",
            "--template",
            str(pptx_template),
            "--config",
            str(_config_file(tmp_path, {"text": {"title": "Hi"}})),
            "--out",
            str(out),
        ]
    )
    assert code == 0
    assert out.is_file()


def test_fill_reports_problems_with_exit_1(pptx_template: Path, tmp_path: Path) -> None:
    code = main(
        [
            "fill",
            "--template",
            str(pptx_template),
            "--config",
            str(_config_file(tmp_path, {"text": {"ghost": "x"}})),
            "--out",
            str(tmp_path / "out.pptx"),
        ]
    )
    assert code == 1


def test_fill_bad_config_exits_2(pptx_template: Path, tmp_path: Path) -> None:
    code = main(
        [
            "fill",
            "--template",
            str(pptx_template),
            "--config",
            str(tmp_path / "missing.json"),
            "--out",
            str(tmp_path / "out.pptx"),
        ]
    )
    assert code == 2


def test_deck_success(tmp_path: Path) -> None:
    out = tmp_path / "deck.pptx"
    code = main(["deck", "--out", str(out), "--complexity", "maximum", "--slides", "5"])
    assert code == 0
    assert len(Presentation(str(out)).slides) == 5


def test_deck_invalid_slides_exits_2(tmp_path: Path) -> None:
    code = main(["deck", "--out", str(tmp_path / "d.pptx"), "--slides", "0"])
    assert code == 2


def test_doc_success(tmp_path: Path) -> None:
    from docx import Document

    out = tmp_path / "document.docx"
    code = main(
        ["doc", "--out", str(out), "--complexity", "complex", "--sections", "4"]
    )
    assert code == 0
    h1 = [
        p
        for p in Document(str(out)).paragraphs
        if p.style and p.style.name == "Heading 1"
    ]
    assert len(h1) == 4


def test_doc_invalid_sections_exits_2(tmp_path: Path) -> None:
    code = main(["doc", "--out", str(tmp_path / "d.docx"), "--sections", "0"])
    assert code == 2


def test_sheet_success(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    out = tmp_path / "workbook.xlsx"
    code = main(
        ["sheet", "--out", str(out), "--complexity", "complex", "--sheets", "3"]
    )
    assert code == 0
    assert len(load_workbook(str(out)).sheetnames) == 3


def test_sheet_invalid_sheets_exits_2(tmp_path: Path) -> None:
    code = main(["sheet", "--out", str(tmp_path / "w.xlsx"), "--sheets", "0"])
    assert code == 2


def test_pdf_success(tmp_path: Path) -> None:
    out = tmp_path / "document.pdf"
    code = main(
        ["pdf", "--out", str(out), "--complexity", "complex", "--sections", "4"]
    )
    assert code == 0
    assert out.read_bytes()[:4] == b"%PDF"


def test_pdf_invalid_sections_exits_2(tmp_path: Path) -> None:
    code = main(["pdf", "--out", str(tmp_path / "d.pdf"), "--sections", "0"])
    assert code == 2


def test_markdown_success(tmp_path: Path) -> None:
    import re

    out = tmp_path / "document.md"
    code = main(
        ["markdown", "--out", str(out), "--complexity", "complex", "--sections", "4"]
    )
    assert code == 0
    text = out.read_text(encoding="utf-8")
    assert text.startswith("# ")
    # Count section headings ("## N. ...") without matching "### " subheadings.
    assert len(re.findall(r"^## \d+\. ", text, re.MULTILINE)) == 4


def test_markdown_md_alias(tmp_path: Path) -> None:
    out = tmp_path / "document.md"
    code = main(["md", "--out", str(out), "--sections", "3"])
    assert code == 0
    assert out.read_text(encoding="utf-8").startswith("# ")


def test_markdown_invalid_sections_exits_2(tmp_path: Path) -> None:
    code = main(["markdown", "--out", str(tmp_path / "d.md"), "--sections", "0"])
    assert code == 2
